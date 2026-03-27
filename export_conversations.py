#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Any
from urllib.parse import quote

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_BASE_URL = "https://api.elevenlabs.io"
LIST_PATH = "/v1/convai/conversations"
DEFAULT_STATE_FILENAME = ".elevenlabs_captured_ids.json"


def flatten_record(obj: Any, prefix: str = "") -> dict[str, Any]:
    out: dict[str, Any] = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            out.update(flatten_record(v, key))
    elif isinstance(obj, list):
        out[prefix] = json.dumps(obj, ensure_ascii=False)
    else:
        out[prefix] = obj
    return out


def sanitize_filename_part(s: str) -> str:
    return re.sub(r'[^\w.\-]+', "_", s.strip())[:120] or "agent"


def load_captured_state(path: str) -> dict[str, set[str]]:
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError:
        print(
            f"Warning: corrupt or invalid JSON in state file {path!r}; starting with empty cache.",
            file=sys.stderr,
        )
        return {}
    except OSError:
        return {}
    raw = data.get("captured") if isinstance(data, dict) else None
    if not isinstance(raw, dict):
        return {}
    out: dict[str, set[str]] = {}
    for agent_id, ids in raw.items():
        if isinstance(ids, list):
            out[str(agent_id)] = {str(x) for x in ids if x is not None}
    return out


def save_captured_state(path: str, state: dict[str, set[str]]) -> None:
    captured = {k: sorted(v) for k, v in sorted(state.items())}
    payload = {"version": 1, "captured": captured}
    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, path)


def prompt_agent_ids() -> list[str]:
    print(
        "Enter one or more agent IDs (comma-separated). "
        "Each agent gets its own Excel file.",
        file=sys.stderr,
    )
    raw = input("Agent ID(s): ").strip()
    if not raw:
        print("No agent IDs entered.", file=sys.stderr)
        sys.exit(1)
    ids = [p.strip() for p in raw.split(",") if p.strip()]
    if not ids:
        print("No valid agent IDs after parsing.", file=sys.stderr)
        sys.exit(1)
    return ids


def list_all_conversations(
    session: requests.Session,
    base_url: str,
    agent_id: str,
    page_size: int,
    pause_sec: float,
    max_pages: int,
) -> list[dict[str, Any]]:
    cursor: str | None = None
    summaries: list[dict[str, Any]] = []
    pages = 0
    while True:
        pages += 1
        params: dict[str, Any] = {"agent_id": agent_id, "page_size": page_size}
        if cursor:
            params["cursor"] = cursor
        url = base_url.rstrip("/") + LIST_PATH
        r = session.get(url, params=params, timeout=120)
        r.raise_for_status()
        data = r.json()
        batch = data.get("conversations") or []
        summaries.extend(batch)
        if not data.get("has_more"):
            break
        if max_pages > 0 and pages >= max_pages:
            if data.get("has_more"):
                print(
                    f"  Stopped list pagination after {max_pages} page(s) (--max-pages); "
                    "more conversations exist on the server.",
                    file=sys.stderr,
                )
            break
        cursor = data.get("next_cursor")
        if not cursor:
            break
        if pause_sec > 0:
            time.sleep(pause_sec)
    return summaries


def get_conversation(
    session: requests.Session,
    base_url: str,
    conversation_id: str,
) -> dict[str, Any]:
    seg = quote(str(conversation_id), safe="")
    url = base_url.rstrip("/") + f"/v1/convai/conversations/{seg}"
    r = session.get(url, timeout=120)
    r.raise_for_status()
    return r.json()


def style_sheet(ws) -> None:
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in column:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def export_agent(
    session: requests.Session,
    base_url: str,
    agent_id: str,
    output_dir: str,
    page_size: int,
    pause_sec: float,
    captured_state: dict[str, set[str]],
    state_path: str | None,
    force_refresh: bool,
    max_pages: int,
) -> str:
    print(f"Listing conversations for agent {agent_id!r}...", file=sys.stderr)
    summaries = list_all_conversations(
        session, base_url, agent_id, page_size, pause_sec, max_pages
    )
    if not summaries:
        print(f"  No conversations found for agent {agent_id!r}.", file=sys.stderr)
    seen_ids: set[str] = set()
    deduped: list[dict[str, Any]] = []
    for s in summaries:
        cid = s.get("conversation_id")
        if not cid:
            continue
        cid = str(cid)
        if cid in seen_ids:
            continue
        seen_ids.add(cid)
        deduped.append(s)
    summaries = deduped

    already = captured_state.setdefault(agent_id, set())
    rows: list[dict[str, Any]] = []
    skipped = 0
    fetched = 0
    for i, s in enumerate(summaries):
        cid = s.get("conversation_id")
        if not cid:
            continue
        cid = str(cid)
        if not force_refresh and cid in already:
            skipped += 1
            print(
                f"  Skip {i + 1}/{len(summaries)} {cid!r} (already captured; no GET)",
                file=sys.stderr,
            )
            continue
        print(
            f"  Fetching {i + 1}/{len(summaries)} conversation {cid!r}...",
            file=sys.stderr,
        )
        detail = get_conversation(session, base_url, cid)
        rows.append(flatten_record(detail))
        already.add(cid)
        fetched += 1
        if pause_sec > 0:
            time.sleep(pause_sec)

    if state_path and fetched > 0:
        save_captured_state(state_path, captured_state)
        print(
            f"  Saved capture state ({len(already)} id(s) for this agent) -> {state_path}",
            file=sys.stderr,
        )
    if skipped:
        print(
            f"  Skipped {skipped} conversation(s) already in state (saved API calls).",
            file=sys.stderr,
        )

    df = pd.DataFrame(rows)
    df = df.sort_index(axis=1)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_agent = sanitize_filename_part(agent_id)
    filename = f"conversations_{safe_agent}_{ts}.xlsx"
    path = os.path.join(output_dir, filename)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conversations")
        style_sheet(writer.sheets["Conversations"])

    print(f"  Wrote {len(rows)} row(s) -> {path}", file=sys.stderr)
    return path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export ElevenLabs ConvAI conversations to Excel (one file per agent)."
    )
    parser.add_argument(
        "--agent-id",
        action="append",
        dest="agent_ids",
        metavar="ID",
        help="Agent ID (repeat for multiple). If omitted, you are prompted.",
    )
    parser.add_argument(
        "--output-dir",
        "-o",
        default=".",
        help="Directory for Excel files (default: current directory).",
    )
    parser.add_argument(
        "--base-url",
        default=os.environ.get("ELEVENLABS_BASE_URL", DEFAULT_BASE_URL),
        help="API base URL (default: %(default)s or ELEVENLABS_BASE_URL).",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=30,
        help="Page size for list endpoint (max 100, default 30).",
    )
    parser.add_argument(
        "--pause",
        type=float,
        default=0.2,
        help="Seconds to sleep between paginated list / detail calls (default 0.2).",
    )
    parser.add_argument(
        "--state-file",
        default=None,
        metavar="PATH",
        help=(
            "JSON file tracking conversation IDs already fetched (skips GET next time). "
            f"Default: {DEFAULT_STATE_FILENAME} under --output-dir."
        ),
    )
    parser.add_argument(
        "--no-state",
        action="store_true",
        help="Do not read or write capture state; every run calls GET for all listed IDs.",
    )
    parser.add_argument(
        "--force-refresh",
        action="store_true",
        help="Ignore capture state and GET every conversation again (state file still updated).",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=0,
        metavar="N",
        help=(
            "Maximum list API pages per agent (each page is up to --page-size conversations). "
            "0 means no limit (default)."
        ),
    )
    args = parser.parse_args()

    api_key = os.environ.get("ELEVENLABS_API_KEY") or os.environ.get("XI_API_KEY")
    if not api_key:
        print(
            "Set ELEVENLABS_API_KEY (or XI_API_KEY) in the environment.",
            file=sys.stderr,
        )
        sys.exit(1)

    agent_ids = args.agent_ids if args.agent_ids else prompt_agent_ids()
    agent_ids = [a.strip() for a in agent_ids if a.strip()]
    if not agent_ids:
        print("No agent IDs after stripping.", file=sys.stderr)
        sys.exit(1)

    if args.page_size < 1 or args.page_size > 100:
        print("--page-size must be between 1 and 100.", file=sys.stderr)
        sys.exit(1)

    if args.max_pages < 0:
        print("--max-pages must be 0 or positive.", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)

    state_path: str | None = None
    captured_state: dict[str, set[str]] = {}
    if not args.no_state:
        state_path = args.state_file or os.path.join(
            args.output_dir, DEFAULT_STATE_FILENAME
        )
        captured_state = load_captured_state(state_path)
        if state_path:
            n = sum(len(v) for v in captured_state.values())
            print(
                f"Loaded capture state: {len(captured_state)} agent(s), {n} conversation id(s) -> {state_path!r}",
                file=sys.stderr,
            )

    session = requests.Session()
    session.headers.update(
        {
            "xi-api-key": api_key,
            "Accept": "application/json",
        }
    )

    for agent_id in agent_ids:
        export_agent(
            session,
            args.base_url,
            agent_id,
            args.output_dir,
            args.page_size,
            args.pause,
            captured_state,
            state_path if not args.no_state else None,
            args.force_refresh,
            args.max_pages,
        )


if __name__ == "__main__":
    main()
